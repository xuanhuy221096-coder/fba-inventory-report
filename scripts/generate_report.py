#!/usr/bin/env python3
"""
FBA Inventory Health Report Generator

Usage:
    python generate_report.py <fba_inventory.csv> <business_report.csv> [options]

Options:
    --prep-days N        Production + labeling days (default: 6)
    --sea-days N         Sea freight days (default: 30)
    --air-days N         Air freight days (default: 15)
    --checkin-days N     Amazon check-in days (default: 7)
    --safety-days N      Safety stock days (default: 14)
    --target-days N      Target cover days after arrival (default: 60)
    --min-order N        Minimum order quantity (default: 50)
    --output PATH        Output file path (default: /mnt/user-data/outputs/Inventory_Health_Report_YYYY-MM-DD.xlsx)
"""

import pandas as pd
import re
import sys
import argparse
import warnings
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# === STYLING CONSTANTS ===
THIN_BORDER = Border(
    left=Side(style='thin', color='D5D8DC'), right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'), bottom=Side(style='thin', color='D5D8DC')
)
HEADER_FILL = PatternFill('solid', fgColor='2C3E50')
HEADER_FONT = Font(bold=True, size=9, color='FFFFFF', name='Arial')
NORMAL_FONT = Font(size=9, name='Arial')
BOLD_FONT = Font(bold=True, size=9, name='Arial')

STATUS_ROW_COLORS = {
    '🔴 DEAD STOCK': 'FADBD8', '⚫ OOS': 'D5D8DC', '🔵 LOW STOCK': 'D6EAF8',
    '🟡 SLOW - Clear': 'FDEBD0', '🟡 SLOW': 'FEF9E7', '🟢 HEALTHY': 'D5F5E3',
}
URGENCY_COLORS = {
    '🚨 AIR URGENT': ('F5B7B1', Font(bold=True, size=10, color='C0392B', name='Arial')),
    '✈️ Air': ('D6EAF8', Font(bold=True, size=10, color='2471A3', name='Arial')),
    '🚢 Sea': ('D5F5E3', Font(bold=True, size=10, color='1E8449', name='Arial')),
}


def identify_files(file1, file2):
    """Auto-detect which file is FBA Inventory vs Business Report."""
    df1 = pd.read_csv(file1, nrows=2)
    df2 = pd.read_csv(file2, nrows=2)

    inv_markers = ['snapshot-date', 'fnsku', 'inv-age-0-to-90-days', 'units-shipped-t7']
    biz_markers = ['(Parent) ASIN', '(Child) ASIN', 'Sessions - Total', 'Unit Session Percentage']

    df1_is_inv = sum(1 for m in inv_markers if m in df1.columns) >= 2
    df2_is_inv = sum(1 for m in inv_markers if m in df2.columns) >= 2
    df1_is_biz = sum(1 for m in biz_markers if m in df1.columns) >= 2
    df2_is_biz = sum(1 for m in biz_markers if m in df2.columns) >= 2

    if df1_is_inv and df2_is_biz:
        return file1, file2
    elif df2_is_inv and df1_is_biz:
        return file2, file1
    else:
        raise ValueError("Cannot auto-detect file types. Ensure one is FBA Inventory and one is Business Report by Child ASIN.")


def load_and_clean(inv_path, biz_path):
    inv = pd.read_csv(inv_path)
    biz = pd.read_csv(biz_path)

    # Validate Business Report is by Child ASIN (not aggregated by date)
    if '(Child) ASIN' not in biz.columns:
        raise ValueError("Business Report must be 'Detail Page Sales and Traffic by Child Item'. Re-download from Seller Central.")

    # Clean Business Report
    for c in ['Units Ordered', 'Sessions - Total', 'Page Views - Total']:
        if c in biz.columns:
            biz[c] = pd.to_numeric(biz[c].astype(str).str.replace(',', ''), errors='coerce').fillna(0).astype(int)
    if 'Ordered Product Sales' in biz.columns:
        biz['Ordered Product Sales'] = pd.to_numeric(biz['Ordered Product Sales'].astype(str).str.replace(r'[$,]', '', regex=True), errors='coerce').fillna(0)
    if 'Unit Session Percentage' in biz.columns:
        biz['Unit Session Percentage'] = pd.to_numeric(biz['Unit Session Percentage'].astype(str).str.replace('%', ''), errors='coerce').fillna(0)

    # Merge
    merged = inv.merge(biz, left_on='asin', right_on='(Child) ASIN', how='left', suffixes=('_inv', '_biz'))

    # Clean FBA Inventory numerics
    num_cols = [
        'available', 'inbound-quantity', 'units-shipped-t7', 'units-shipped-t30', 'units-shipped-t60', 'units-shipped-t90',
        'estimated-excess-quantity', 'sell-through', 'weeks-of-cover-t30', 'weeks-of-cover-t90', 'sales-rank',
        'inv-age-0-to-90-days', 'inv-age-91-to-180-days', 'inv-age-181-to-270-days', 'inv-age-271-to-365-days',
        'inv-age-366-to-455-days', 'inv-age-456-plus-days', 'estimated-storage-cost-next-month',
        'Total Reserved Quantity', 'unfulfillable-quantity',
        'Reserved FC Transfer', 'Reserved Customer Order', 'Reserved FC Processing', 'Reserved Staging',
        'sales-shipped-last-7-days', 'sales-shipped-last-30-days', 'sales-shipped-last-60-days', 'sales-shipped-last-90-days',
        'your-price', 'inbound-working', 'inbound-shipped', 'inbound-received'
    ]
    for c in num_cols:
        if c in merged.columns:
            merged[c] = pd.to_numeric(merged[c], errors='coerce').fillna(0)

    return merged


def parse_model_color(name):
    if pd.isna(name):
        return ('Unknown', 'N/A')
    name = str(name)

    # Model detection (order matters: specific before general).
    # NOTE: 1500-series checks must run BEFORE their 2500/3500 counterparts,
    # because some 1500 titles contain "Not for Ram 2500/3500" / similar
    # exclusion text that would otherwise be misclassified.
    model = 'Other'

    if any(k in name for k in ['Super Duty', 'F250', 'F350', 'F450']):
        model = 'F250/F350/F450 Super Duty'
    elif 'F150' in name or 'F-150' in name:
        if 'Bronco' in name and 'Tundra' in name:
            model = 'F150 / Bronco / Tundra'
        elif 'Bronco' in name:
            model = 'F150 / Bronco'
        else:
            model = 'F150'
    elif 'Extended Tow Hook' in name and 'Silverado 1500' in name:
        model = 'Silverado 1500 Extended'
    elif 'Silverado 1500' in name:
        model = 'Silverado 1500'
    elif 'Silverado 2500' in name or 'Silverado 3500' in name:
        model = 'Silverado 2500/3500 HD'
    elif 'Sierra 1500' in name:
        model = 'Sierra 1500'
    elif 'Sierra 2500' in name or 'Sierra 3500' in name:
        model = 'Sierra 2500/3500 HD'
    elif 'Ram 1500' in name:
        model = 'Ram 1500'
    elif 'Ram 2500' in name or 'Ram 3500' in name:
        model = 'Ram 2500/3500'
    elif 'Colorado' in name or 'Canyon' in name:
        model = 'Colorado / Canyon'
    elif 'Tire Valve' in name:
        model = 'Tire Valve Stem Caps'

    # Color detection
    color = 'N/A'
    color_match = re.search(r'\(([^)]+)\)\s*$', name)
    if color_match:
        color = color_match.group(1).strip()
    else:
        color_match2 = re.search(r',\s*(\w[\w\s]*?)\s*$', name)
        if color_match2:
            candidate = color_match2.group(1).strip()
            known = ['Red', 'Blue', 'Black', 'White', 'Yellow', 'Orange', 'Green', 'Purple', 'Pink', 'American', 'Camo', 'Gray', 'Silver']
            if candidate in known:
                color = candidate

    return (model, color)


def _blend_velocity(row):
    v7 = row['units-shipped-t7'] / 7
    v30 = row['units-shipped-t30'] / 30
    v90 = row['units-shipped-t90'] / 90
    return 0.5 * v30 + 0.3 * v7 + 0.2 * v90

def _velocity_cv(row):
    import statistics
    v7 = row['units-shipped-t7'] / 7
    v30 = row['units-shipped-t30'] / 30
    v90 = row['units-shipped-t90'] / 90
    vb = 0.5 * v30 + 0.3 * v7 + 0.2 * v90
    if vb <= 0:
        return 0.0
    return statistics.pstdev([v7, v30, v90]) / vb

def compute_metrics(merged):
    merged['daily_velocity'] = merged.apply(_blend_velocity, axis=1)
    merged['vel_cv'] = merged.apply(_velocity_cv, axis=1)
    # On-hand = available + the 4 reserved sub-columns. Amazon's "Total Reserved
    # Quantity" column is unreliable (often undercounts FC Transfer), so sum the parts.
    for c in ['Reserved FC Transfer', 'Reserved Customer Order', 'Reserved FC Processing', 'Reserved Staging']:
        if c not in merged.columns:
            merged[c] = 0
        merged[c] = pd.to_numeric(merged[c], errors='coerce').fillna(0)
    merged['reserved_total'] = (merged['Reserved FC Transfer'] + merged['Reserved Customer Order']
                                + merged['Reserved FC Processing'] + merged['Reserved Staging'])
    merged['total_fba'] = merged['available'] + merged['reserved_total'] + merged['unfulfillable-quantity']
    merged['days_remaining'] = merged.apply(lambda r: round(r['total_fba'] / r['daily_velocity']) if r['daily_velocity'] > 0 else 9999, axis=1)
    merged['aged_181plus'] = (merged['inv-age-181-to-270-days'] + merged['inv-age-271-to-365-days'] +
                              merged['inv-age-366-to-455-days'] + merged['inv-age-456-plus-days'])
    return merged


def assign_status(row):
    # ORDER MATTERS: OOS and LOW STOCK are checked BEFORE SLOW.
    # A SKU that is out of stock can still show a high days_remaining artifact,
    # so the stockout/low checks must win over the slow-mover checks.
    if row.get('no-sale-last-6-months') == 'Y' or (row['units-shipped-t90'] == 0 and row['total_fba'] > 0):
        return '🔴 DEAD STOCK'
    if row['total_fba'] <= 0 and row['inbound-quantity'] <= 0:
        return '⚫ OOS'
    if row['days_remaining'] < 14:
        return '🔵 LOW STOCK'
    if row['days_remaining'] > 180:
        return '🟡 SLOW - Clear'
    if row['days_remaining'] > 90:
        return '🟡 SLOW'
    return '🟢 HEALTHY'


def compute_restock(row, config):
    vel = row['daily_velocity']
    total_supply = row['total_fba'] + row['inbound-quantity']

    if row['status'] == '🔴 DEAD STOCK':
        return ('❌ No Restock', 'N/A', 0, 'Dead stock')
    if vel <= 0 and row['units-shipped-t90'] <= 5:
        return ('❌ No Restock', 'N/A', 0, 'Too slow')
    if vel <= 0 and row['units-shipped-t90'] > 5:
        vel = row['units-shipped-t90'] / 90

    # Dynamic safety stock: more buffer for erratic SKUs (high CV)
    safety_units = vel * config['safety'] * (1 + row.get('vel_cv', 0))

    total_days = total_supply / vel if vel > 0 else 9999
    sea_total = config['prep'] + config['sea'] + config['checkin']
    air_total = config['prep'] + config['air'] + config['checkin']

    if total_days > (sea_total + config['safety']):
        qty = max(0, round((sea_total + config['target']) * vel + safety_units - total_supply))
        if qty <= 0:
            return ('✅ OK', 'N/A', 0, f'Covered {int(total_days)}d')
        qty = max(qty, config['min_order'])
        return ('🚢 Sea', 'Sea Freight', qty, f'{int(total_days)}d left → order {qty} by sea')
    elif total_days > (air_total + config['safety']):
        qty = max(0, round((sea_total + config['target']) * vel + safety_units - total_supply))
        qty = max(qty, config['min_order'])
        return ('✈️ Air', 'Air Freight', qty, f'{int(total_days)}d left → air {qty} units')
    else:
        qty = max(0, round((air_total + config['target']) * vel + safety_units - total_supply))
        qty = max(qty, config['min_order'])
        return ('🚨 AIR URGENT', 'Air Freight ASAP', qty, f'⚠️ {int(total_days)}d left → rush air {qty} units')


def build_dashboard(wb, merged, config):
    ws = wb.active
    ws.title = "Dashboard"

    ws.merge_cells('A1:H1')
    ws['A1'] = '📦 INVENTORY HEALTH REPORT'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1A5276')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:H2')
    snapshot = merged['snapshot-date'].iloc[0] if 'snapshot-date' in merged.columns else 'N/A'
    ws['A2'] = f'Snapshot: {snapshot} | Total SKUs: {len(merged)} | Available: {int(merged["available"].sum())} | Inbound: {int(merged["inbound-quantity"].sum())}'
    ws['A2'].font = Font(size=10, color='566573')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Lead time assumptions
    ws.merge_cells('A4:H4')
    ws['A4'] = '📋 LEAD TIME ASSUMPTIONS'
    ws['A4'].font = Font(bold=True, size=11, color='FFFFFF')
    ws['A4'].fill = PatternFill('solid', fgColor='2980B9')

    sea_total = config['prep'] + config['sea'] + config['checkin']
    air_total = config['prep'] + config['air'] + config['checkin']
    lt_data = [
        ('Production + Labeling', f'{config["prep"]}d'),
        ('Sea Freight', f'{config["sea"]}d'),
        ('Air Freight', f'{config["air"]}d'),
        ('Amazon Check-in', f'{config["checkin"]}d'),
        ('Total Lead (Sea)', f'{sea_total}d'),
        ('Total Lead (Air)', f'{air_total}d'),
        ('Safety Stock', f'{config["safety"]}d'),
        ('Restock Target', f'{config["target"]}d cover'),
        ('Min Order Qty', f'{config["min_order"]} units'),
    ]
    for i, (label, val) in enumerate(lt_data):
        ws.cell(row=5 + i, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=5 + i, column=2, value=val).font = Font(bold=True, size=10, name='Arial', color='1A5276')

    # Metrics
    r = 15
    metrics = [
        ('Total Revenue 30d', f'${merged["sales-shipped-last-30-days"].sum():,.0f}'),
        ('Units Shipped 30d', f'{int(merged["units-shipped-t30"].sum())}'),
        ('Total Daily Velocity', f'{merged["daily_velocity"].sum():.1f} units/day'),
        ('Est. Storage Cost', f'${merged["estimated-storage-cost-next-month"].sum():,.2f}'),
    ]
    for i, (label, val) in enumerate(metrics):
        col = i * 2 + 1
        ws.cell(row=r, column=col, value=label).font = Font(bold=True, size=9, color='566573')
        ws.cell(row=r + 1, column=col, value=val).font = Font(bold=True, size=12, color='1A5276')

    # Status breakdown
    r = 18
    ws.cell(row=r, column=1, value='STATUS BREAKDOWN').font = Font(bold=True, size=11, color='2C3E50')
    for st, cnt in merged['status'].value_counts().items():
        r += 1
        ws.cell(row=r, column=1, value=st).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=cnt).font = BOLD_FONT

    # Restock summary
    r += 2
    ws.cell(row=r, column=1, value='RESTOCK SUMMARY').font = Font(bold=True, size=11, color='2C3E50')
    for act, cnt in merged['Restock Action'].value_counts().items():
        r += 1
        ws.cell(row=r, column=1, value=act).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=cnt).font = BOLD_FONT

    # Revenue by model
    r += 2
    ws.cell(row=r, column=1, value='REVENUE BY MODEL (30d)').font = Font(bold=True, size=11, color='2C3E50')
    for model, rev in merged.groupby('Model')['sales-shipped-last-30-days'].sum().sort_values(ascending=False).items():
        r += 1
        ws.cell(row=r, column=1, value=model).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=f'${rev:,.0f}').font = BOLD_FONT


def build_restock_order(wb, merged, config):
    restock = merged[merged['Restock Qty'] > 0].copy()
    if len(restock) == 0:
        return

    ws = wb.create_sheet("📦 Restock Order", 1)

    ws.merge_cells('A1:M1')
    ws['A1'] = '📦 RESTOCK ORDER LIST'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1A5276')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    air_urgent = restock[restock['Restock Action'] == '🚨 AIR URGENT']
    air_normal = restock[restock['Restock Action'] == '✈️ Air']
    sea = restock[restock['Restock Action'] == '🚢 Sea']

    ws.merge_cells('A2:M2')
    ws['A2'] = f'🚨 Air Urgent: {len(air_urgent)} SKUs ({int(air_urgent["Restock Qty"].sum())} units) | ✈️ Air: {len(air_normal)} SKUs ({int(air_normal["Restock Qty"].sum())} units) | 🚢 Sea: {len(sea)} SKUs ({int(sea["Restock Qty"].sum())} units) | TOTAL: {int(restock["Restock Qty"].sum())} units'
    ws['A2'].font = Font(size=10, color='566573')
    ws['A2'].alignment = Alignment(horizontal='center')

    sea_total = config['prep'] + config['sea'] + config['checkin']
    air_total = config['prep'] + config['air'] + config['checkin']
    ws.merge_cells('A3:M3')
    ws['A3'] = f'Min order: {config["min_order"]} units | Lead: Prep {config["prep"]}d + Sea {config["sea"]}d / Air {config["air"]}d + Check-in {config["checkin"]}d | Safety: {config["safety"]}d | Target: {config["target"]}d cover'
    ws['A3'].font = Font(size=9, color='7F8C8D', italic=True)

    headers = ['#', 'Urgency', 'Model', 'Color', 'ASIN', 'Ship Method', 'Order Qty', 'Available', 'Inbound', 'Total Supply', 'Daily Vel.', 'Days Left', 'Note']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[5].height = 28
    ws.auto_filter.ref = f'A5:{get_column_letter(len(headers))}5'

    row_idx = 6
    item_num = 1
    for action_type in ['🚨 AIR URGENT', '✈️ Air', '🚢 Sea']:
        subset = restock[restock['Restock Action'] == action_type]
        for _, r in subset.iterrows():
            fill_color, urgency_font = URGENCY_COLORS[action_type]
            total_supply = int(r['total_fba'] + r['inbound-quantity'])
            data = [item_num, r['Restock Action'], r['Model'], r['Color'], r['asin'],
                    r['Ship Method'], int(r['Restock Qty']), int(r['available']),
                    int(r['inbound-quantity']), total_supply,
                    round(r['daily_velocity'], 1),
                    int(r['days_remaining']) if r['days_remaining'] < 9999 else 'N/A',
                    r['Restock Note']]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.fill = PatternFill('solid', fgColor=fill_color)
                if col_idx in [1, 7, 8, 9, 10, 11, 12]:
                    cell.alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=2).font = urgency_font
            ws.cell(row=row_idx, column=7).font = Font(bold=True, size=11, name='Arial')
            row_idx += 1
            item_num += 1

    # Total row
    row_idx += 1
    ws.cell(row=row_idx, column=5, value='TOTAL').font = Font(bold=True, size=11, name='Arial')
    ws.cell(row=row_idx, column=7, value=int(restock['Restock Qty'].sum())).font = Font(bold=True, size=12, color='C0392B', name='Arial')
    ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='center')
    for c in range(1, len(headers) + 1):
        ws.cell(row=row_idx, column=c).fill = PatternFill('solid', fgColor='F7DC6F')
        ws.cell(row=row_idx, column=c).border = THIN_BORDER

    widths = [5, 18, 28, 18, 16, 18, 10, 10, 10, 12, 10, 10, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A6'


def build_inventory_detail(wb, merged):
    ws = wb.create_sheet("Inventory Detail")

    headers = ['Status', 'Model', 'Color', 'ASIN', 'Price',
               'Available', 'Reserved', 'Inbound', 'Total FBA',
               'Sold 7d', 'Sold 30d', 'Sold 90d', 'Daily Vel.', 'Days Left', 'Sell-Through',
               'Sessions', 'Units Ordered', 'CVR%', 'Revenue 30d',
               'Age 0-90', 'Age 91-180', 'Age 181+', 'Storage Cost', 'Sales Rank', 'Health',
               'Restock Action', 'Ship Method', 'Restock Qty', 'Restock Note']

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill('solid', fgColor='E74C3C' if col_idx >= 26 else '2C3E50')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    sorted_df = merged.sort_values(['Model', 'Color'])
    for row_idx, (_, r) in enumerate(sorted_df.iterrows(), 2):
        data = [
            r['status'], r['Model'], r['Color'], r['asin'], r['your-price'],
            int(r['available']), int(r['reserved_total']), int(r['inbound-quantity']), int(r['total_fba']),
            int(r['units-shipped-t7']), int(r['units-shipped-t30']), int(r['units-shipped-t90']),
            round(r['daily_velocity'], 1), int(r['days_remaining']) if r['days_remaining'] < 9999 else 'N/A',
            r['sell-through'],
            r.get('Sessions - Total', 0), r.get('Units Ordered', 0), r.get('Unit Session Percentage', 0),
            round(r['sales-shipped-last-30-days'], 2),
            int(r['inv-age-0-to-90-days']), int(r['inv-age-91-to-180-days']), int(r['aged_181plus']),
            round(r['estimated-storage-cost-next-month'], 2),
            int(r['sales-rank']) if r['sales-rank'] > 0 else '',
            r.get('fba-inventory-level-health-status', ''),
            r['Restock Action'], r['Ship Method'], int(r['Restock Qty']), r['Restock Note']
        ]
        fill_color = STATUS_ROW_COLORS.get(r['status'], 'FFFFFF')
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.fill = PatternFill('solid', fgColor=fill_color)
            if col_idx in [5, 19, 23]:
                cell.number_format = '$#,##0.00'
            if col_idx in [6, 7, 8, 9, 10, 11, 12, 13, 20, 21, 22, 28]:
                cell.alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=26).font = BOLD_FONT

    widths = [18, 28, 16, 14, 8, 8, 8, 8, 8, 7, 7, 7, 7, 8, 8, 8, 8, 6, 10, 8, 8, 8, 8, 8, 10, 16, 16, 10, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'


def build_action_items(wb, merged):
    ws = wb.create_sheet("⚠️ Action Items")

    ws.merge_cells('A1:H1')
    ws['A1'] = '⚠️ ITEMS REQUIRING ACTION'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='E74C3C')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Priority', 'Model', 'Color', 'ASIN', 'Issue', 'Available', 'Ship Method', 'Recommendation']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, color='FFFFFF', name='Arial')
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    action_row = 4

    def write_row(row_num, vals, fill_color):
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.font = NORMAL_FONT
            cell.fill = PatternFill('solid', fgColor=fill_color)
            cell.border = THIN_BORDER
        return row_num + 1

    # Dead stock
    for _, r in merged[merged['status'] == '🔴 DEAD STOCK'].iterrows():
        vals = ['🔴 URGENT', r['Model'], r['Color'], r['asin'], 'Dead stock - 0 sales 90d', int(r['available']), 'N/A', 'Remove / Liquidate']
        action_row = write_row(action_row, vals, 'FADBD8')

    # OOS
    for _, r in merged[merged['status'] == '⚫ OOS'].iterrows():
        vals = ['⚫ URGENT', r['Model'], r['Color'], r['asin'], f'OOS - selling {int(r["units-shipped-t30"])}/mo', 0, r['Ship Method'], r['Restock Note']]
        action_row = write_row(action_row, vals, 'D5D8DC')

    # Low stock
    for _, r in merged[merged['status'] == '🔵 LOW STOCK'].iterrows():
        vals = ['🔵 HIGH', r['Model'], r['Color'], r['asin'], f'Low stock - {int(r["days_remaining"])}d @ {r["daily_velocity"]:.1f}/d',
                int(r['available']), r['Ship Method'], r['Restock Note']]
        action_row = write_row(action_row, vals, 'D6EAF8')

    # Air urgent
    for _, r in merged[merged['Restock Action'] == '🚨 AIR URGENT'].iterrows():
        if r['status'] in ['🔴 DEAD STOCK', '⚫ OOS', '🔵 LOW STOCK']:
            continue
        vals = ['🚨 RESTOCK', r['Model'], r['Color'], r['asin'],
                f'{int(r["days_remaining"])}d left, vel={r["daily_velocity"]:.1f}/d',
                int(r['available']), 'AIR URGENT', f'Rush air {int(r["Restock Qty"])} units']
        action_row = write_row(action_row, vals, 'F5B7B1')

    # Slow movers
    for _, r in merged[merged['status'].str.contains('SLOW')].sort_values('days_remaining', ascending=False).iterrows():
        vals = ['🟡 SLOW', r['Model'], r['Color'], r['asin'],
                f'{int(r["days_remaining"])}d supply, {int(r["aged_181plus"])} aged 181+',
                int(r['available']), 'N/A', 'Run Sale / Remove']
        action_row = write_row(action_row, vals, 'FDEBD0')

    widths = [14, 28, 16, 16, 32, 10, 16, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'


def _verify_report(merged):
    """Mandatory self-check. Raises AssertionError if any integrity rule fails.
    Run before saving so a bad report never reaches the user silently."""
    errors = []

    # 1. On-hand identity: total_fba must equal available + reserved_total + unfulfillable
    bad = merged[(merged['total_fba'] - (merged['available'] + merged['reserved_total']
                 + merged['unfulfillable-quantity'])).abs() > 0.5]
    if len(bad):
        errors.append(f"{len(bad)} rows fail total_fba = available + reserved_total + unfulfillable")

    # 2. No negative core quantities
    for col in ['available', 'total_fba', 'inbound-quantity', 'reserved_total']:
        if (merged[col] < 0).any():
            errors.append(f"negative values in {col}")

    # 3. Status sanity: OOS rows must have total_fba <= 0 and no inbound
    oos = merged[merged['status'] == '⚫ OOS']
    if ((oos['total_fba'] > 0) | (oos['inbound-quantity'] > 0)).any():
        errors.append("OOS rows with stock or inbound present")

    # 4. Status sanity: any row with stock and velocity must not be OOS
    has_stock = merged[(merged['total_fba'] > 0) & (merged['daily_velocity'] > 0)]
    if (has_stock['status'] == '⚫ OOS').any():
        errors.append("rows with stock+velocity mislabeled OOS")

    # 5. Every restock row must have a ship method
    rq = merged[merged['Restock Qty'] > 0]
    if (rq['Ship Method'] == 'N/A').any():
        errors.append("restock rows missing ship method")

    if errors:
        raise AssertionError("REPORT VERIFY FAILED:\n  - " + "\n  - ".join(errors))
    print(f"✅ Self-check passed ({len(merged)} SKUs verified)")


def generate_report(inv_path, biz_path, config, output_path=None):
    merged = load_and_clean(inv_path, biz_path)
    merged[['Model', 'Color']] = merged['product-name'].apply(lambda x: pd.Series(parse_model_color(x)))
    merged = compute_metrics(merged)
    merged['status'] = merged.apply(assign_status, axis=1)
    merged[['Restock Action', 'Ship Method', 'Restock Qty', 'Restock Note']] = merged.apply(
        lambda r: pd.Series(compute_restock(r, config)), axis=1)

    _verify_report(merged)

    if output_path is None:
        snapshot = merged['snapshot-date'].iloc[0] if 'snapshot-date' in merged.columns else 'report'
        output_path = f'/mnt/user-data/outputs/Inventory_Health_Report_{snapshot}.xlsx'

    wb = Workbook()
    build_dashboard(wb, merged, config)
    build_restock_order(wb, merged, config)
    build_inventory_detail(wb, merged)
    build_action_items(wb, merged)
    wb.save(output_path)

    # Print summary
    restock = merged[merged['Restock Qty'] > 0]
    air_urgent = restock[restock['Restock Action'] == '🚨 AIR URGENT']
    air_normal = restock[restock['Restock Action'] == '✈️ Air']
    sea = restock[restock['Restock Action'] == '🚢 Sea']

    print(f"✅ Report saved to: {output_path}")
    print(f"\n📦 Inventory Health Report — {merged.get('snapshot-date', pd.Series(['N/A'])).iloc[0]}")
    print(f"Overview: {len(merged)} SKUs | {int(merged['available'].sum())} available | {int(merged['inbound-quantity'].sum())} inbound")
    print(f"\nStatus:")
    for st, cnt in merged['status'].value_counts().items():
        print(f"  {st}: {cnt}")
    print(f"\n📦 Restock Total: {int(restock['Restock Qty'].sum())} units")
    print(f"  🚨 Air Urgent: {len(air_urgent)} SKUs, {int(air_urgent['Restock Qty'].sum())} units")
    print(f"  ✈️ Air: {len(air_normal)} SKUs, {int(air_normal['Restock Qty'].sum())} units")
    print(f"  🚢 Sea: {len(sea)} SKUs, {int(sea['Restock Qty'].sum())} units")

    return output_path


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='FBA Inventory Health Report Generator')
    parser.add_argument('file1', help='First CSV file (auto-detected)')
    parser.add_argument('file2', help='Second CSV file (auto-detected)')
    parser.add_argument('--prep-days', type=int, default=6)
    parser.add_argument('--sea-days', type=int, default=30)
    parser.add_argument('--air-days', type=int, default=15)
    parser.add_argument('--checkin-days', type=int, default=7)
    parser.add_argument('--safety-days', type=int, default=14)
    parser.add_argument('--target-days', type=int, default=60)
    parser.add_argument('--min-order', type=int, default=50)
    parser.add_argument('--output', type=str, default=None)
    args = parser.parse_args()

    config = {
        'prep': args.prep_days, 'sea': args.sea_days, 'air': args.air_days,
        'checkin': args.checkin_days, 'safety': args.safety_days,
        'target': args.target_days, 'min_order': args.min_order,
    }

    inv_path, biz_path = identify_files(args.file1, args.file2)
    generate_report(inv_path, biz_path, config, args.output)
